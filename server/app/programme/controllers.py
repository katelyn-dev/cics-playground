from flask import Blueprint, jsonify, make_response, request, send_file
from sqlalchemy import cast, func, not_, and_
from app import db
from .models import Programme, programme_schema, programmes_schema, Application, Students, student_schema, students_schema, EmergencyContact
from datetime import datetime
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import pandas as pd
from io import BytesIO
import os

mod = Blueprint('programmes', __name__)

# http://127.0.0.1:8080/programme?id=1
"""
    Query parameters:
    - id (optional): The ID of the class to retrieve. Return all classes if ID is not provided
"""
@mod.route('/programme', methods=['GET'])
def getAll():
    if 'id' in request.args:
        programme_id = request.args['id']
        print(programme_id)

        programmes = Programme.query.filter_by(class_group_id=programme_id).all()
        if programmes:
            return make_response(programmes_schema.dump(programmes), 200)
        return make_response(jsonify({'message': 'programme not found'}), 404)

    programmes = Programme.query.all()
    return programmes_schema.dump(programmes)

# http://127.0.0.1:8080/searchProgramme?startDate=yyyy-mm-dd&endDate=yyyy-mm-dd
"""
    Query parameters:
    - startDate (required): The start date for the search in YYYY-MM-DD format.
    - endDate (required): The end date for the search in YYYY-MM-DD format.
"""
@mod.route('/searchProgramme', methods=['GET'])
def searchProgramme():
    programme_id = request.args.get('id')
    name = request.args.get('name')
    start_date_str = request.args.get('startDate')
    end_date_str = request.args.get('endDate')

    # Build query filters dynamically
    filters = []

    filters.append(not_(Programme.class_group_id.startswith("SCL")))
    if programme_id:
        filters.append(Programme.class_group_id == programme_id)

    if name:
        filters.append(Programme.class_name_eng.like(f"%{name}%"))  # Use LIKE for partial matches

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            filters.append(Programme.class_start >= start_date)
        except ValueError:
            return jsonify({"error": "Invalid start date format"}), 400

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            filters.append(Programme.class_end <= end_date)
        except ValueError:
            return jsonify({"error": "Invalid end date format"}), 400

    # Construct and print the query for debugging
    query = Programme.query.filter(*filters)
    print("Constructed Query:", str(query))

    # Execute the query
    programmes = query.all()

    if programmes:
        return make_response(programmes_schema.dump(programmes), 200)
    else:
        return make_response(jsonify({'message': 'Classes not found'}), 404)

# http://127.0.0.1:8080/saveProgramme 
"""
    Example JSON payload:
    {
        "class_group_id": "CL2001",
        "class_name_eng": "Summer 2024 Specific Skills Camp (Age 10-14)",
        "class_name_zhcn": "2024技能夏令营 (10-14 岁)",
        "class_name_zhhk": "2024技能夏令營 (10-14 歲)",
        "target_audience": "junior",
        "class_start": "2024-07-30",
        "class_end": "2024-09-15",
        "extra_attributes": {
            "extra_attributes": [
                "Before Care (7:30am to 9:00am)",
                "After Care (3:00pm to 6:00pm)",
                "Before and After Care (7:30am to 9:00am & 3:00pm to 6:00pm)",
                "Do not need"
            ]
        },
        "extra_attributes_name": "Before and/or After Camp Option",
        "has_extra_attributes": "T",
        "has_subclass": "T",
        "start_time": "2024-07-25T02:54:08",
        "subclass_group_id": "SCL2001"
    }
"""
@mod.route('/saveProgramme', methods=['POST'])
def saveProgramme():
    data = request.get_json()
    
    class_prefix = 'CL'
    max_class_group_id = db.session.query(
        db.func.max(db.cast(db.func.substring(Programme.class_group_id, len(class_prefix) + 2), db.Integer))
    ).filter(Programme.class_group_id.like(f"{class_prefix}_%")).scalar()
    new_class_group = 1 if max_class_group_id is None else max_class_group_id + 1
    new_class_group_id = f"{class_prefix}{new_class_group:05d}"

    if not data:
        return make_response(jsonify({'message': "Invalid request body"}), 400)

    if 'class_name_eng' not in data or 'class_name_zhcn' not in data or 'class_name_zhhk' not in data:
        return make_response(jsonify({'message': "Missing class_name_eng / class_name_zhcn / class_name_zhhk"}), 400)

    class_start = data.get('class_start')
    class_end = data.get('class_end')
    start_time = data.get('start_time')
    if not class_start or not class_end:
        return make_response(jsonify({'message': "Missing class_start / class_end"}), 400)

    try:
        class_start = datetime.strptime(class_start, '%Y-%m-%d').date()
        class_end = datetime.strptime(class_end, '%Y-%m-%d').date()
        if start_time:
             start_time = datetime.strptime(data.get('start_time'), '%Y-%m-%dT%H:%M:%S')
    except ValueError as e:
        return make_response(jsonify({'message': 'Bad Request', 'error': str(e)}), 400)

    if 'extra_attributes' in data and not isinstance(data['extra_attributes'], dict):
        return make_response(jsonify({"error": "Invalid extra_attibutes, expected a dictionary"}), 400)

    new_programme = Programme(class_group_id=new_class_group_id, 
                         class_name_eng=data.get('class_name_eng'),
                         class_name_zhcn=data.get('class_name_zhcn'),
                         class_name_zhhk=data.get('class_name_zhhk'),
                         target_audience=data.get('target_audience'),
                         has_subclass = data.get('has_subclass'),
                         class_fee = data.get('class_fee'),
                         subclass_group_id = data.get('subclass_group_id'),
                         has_extra_attributes = data.get('has_extra_attributes'),
                         extra_attributes_name = data.get('extra_attributes_name'),
                         extra_attributes = data.get('extra_attributes'),
                         class_start=class_start,
                         class_end=class_end,
                         start_time=start_time
                         )

    db.session.add(new_programme)
    db.session.commit()

    return make_response(jsonify(programme_schema.dump(new_programme)),201)

@mod.route("/saveStudent", methods=['POST'])
def save_student():
    data = request.get_json()
    if not data:
        return make_response(jsonify({'message': "Invalid request body"}), 400)
    formId = data.get("formId")
    class_group_id = ''
    from app.form.models import Form
    form = Form.query.filter_by(form_id=formId).first()
    if form:
       class_group_id = form.class_group_id
       print(class_group_id)

    required_fields = ['lastname', 'firstname', 'phone_number', 'email']
    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return make_response(jsonify({'message': f"Missing {', '.join(missing_fields)}"}), 400)
    email = data.get('email')
    lastname = data.get('lastname')
    firstname = data.get('firstname')
    result = check_student_exist(data, class_group_id, email, lastname, firstname)
    return make_response(result,200)

def check_student_exist(data, class_group_id, email, lastname, firstname):
    filters = []
    filters.append(Students.email == email)
    filters.append(Students.lastname == lastname)
    filters.append(Students.firstname == firstname)  
    student_id_query = Students.query.filter(*filters)
    students = student_id_query.all()
    if students:
        student_id = [student.id for student in students]
        n_student_id = student_id[0]
        applications = db.session.query(Application).filter(Application.student_id == n_student_id).all()
        if applications:
            current_application = [application for application in applications][0]
            return { "id" : current_application.id, 
                    "status": "already applied",
                    "class_group_id" : class_group_id}
        else: 
            applicaiton = create_application(data, n_student_id, class_group_id)[0]
            return { "id" : applicaiton.id, 
                    "status": "current student applied",
                    "class_group_id" : class_group_id}
    else:
        new_student_id = create_student(data)
        new_application_id = create_application(data, new_student_id, class_group_id)[0]
        return  { "id" : new_application_id.id, 
                 "status": "new stutend applied", 
                 "class_group_id" : class_group_id}

def create_student(data):
    student_model = Students()
    new_student_id = student_model.create_student(data)
    return new_student_id

def create_application(data, student_id, class_group_id):
    application_model = Application()
    application_id = application_model.create_application(data, student_id, class_group_id)
    return application_id

# Usage: searchStudent?email=chantaiman@email.com&firstname=Tai&lastname=CHAN
@mod.route("/searchStudent", methods=["GET"])
def searchStudent():
    email = request.args.get('email')
    phone_num = request.args.get('phone_num')
    lastname = request.args.get('lastname')
    firstname = request.args.get('firstname')

    # Build query filters dynamically
    filters = []
    if email:
        filters.append(Students.email.like(f"%{email}%")) 
    if phone_num:
        filters.append(Students.phone_number.like(f"%{phone_num}%")) 
    if lastname:
        filters.append(Students.lastname.like(f"%{lastname}%")) 
    if firstname:
        filters.append(Students.firstname.like(f"%{firstname}%")) 

    query = Students.query.filter(*filters)
    print("Constructed Query:", str(query))
    # Execute the query
    students = query.all()

    if students:
        return make_response(students_schema.dump(students), 200)
    else:
        return make_response(jsonify([]), 200)
    

# Usage: /getApplicationDetails --> to get all table records that left join to application
# /getApplicationDetails?filters={"students":{"gender":"M"}}&output_fields=students.firstname&output_fields=programme.class_name_eng
@mod.route('/getApplicationDetails', methods=['GET'])
def getApplicationDetails():
    filters = request.args.get('filters', default="{}")
    output_fields = request.args.getlist('output_fields')

    # Parse filters from JSON string to dictionary
    filters = json.loads(filters)

    # Initialize the query with a left join
    query = db.session.query(Application).outerjoin(Application.student).outerjoin(Application.class_info).outerjoin(Application.emergency_contact_info)

    # Apply filters
    if filters:
        filter_conditions = []
        for table, conditions in filters.items():
            for column, value in conditions.items():
                filter_conditions.append(getattr(globals()[table.capitalize()], column) == value)
        query = query.filter(and_(*filter_conditions))

    # Define the output fields
    if output_fields:
        selected_fields = []
        for field in output_fields:
            table, column = field.split('.')
            selected_fields.append(getattr(globals()[table.capitalize()], column))
        query = query.with_entities(*selected_fields)
    else:
        query = query.with_entities(Application, Students, Programme, EmergencyContact)

    # Fetch the results
    results = query.all()

    # Serialize the results
    output = []
    if output_fields:
        for result in results:
            output.append({field: getattr(result, field.split('.')[-1]) for field in output_fields})
    else:
        for application, student, programme, emergency_contact in results:
            output.append({
                'application': {key: value for key, value in application.__dict__.items() if key != '_sa_instance_state'},
                'student': {key: value for key, value in (student.__dict__ if student else {}).items() if key != '_sa_instance_state'},
                'programme': {key: value for key, value in (programme.__dict__ if programme else {}).items() if key != '_sa_instance_state'},
                'emergency_contact': {key: value for key, value in (emergency_contact.__dict__ if emergency_contact else {}).items() if key != '_sa_instance_state'}
            })

    return jsonify(output)


@mod.route('/export', methods=['GET'])
def export():
    try:
        response = getApplicationDetails()
        data = response.get_json()

        export_data = []
        for item in data:
            programme = item.get('programme', {})
            student = item.get('student', {})
            application = item.get('application', {})
            
            # Transform 'is_paid' value
            is_paid_value = application.get('is_paid', '')
            if is_paid_value == '1':
                is_paid_value = "Yes"
            else:
                is_paid_value = "No"

            # Create the output field: Firstname Lastname/Class_group_id/Phone_number/Email
            output_field = f"{student.get('firstname', '')} {student.get('lastname', '')}/" \
                           f"{application.get('class_group_id', '')}/" \
                           f"{student.get('phone_number', '')}/" \
                           f"{student.get('email', '')}"

            export_data.append({
                'Class Name': programme.get('class_name_eng', ''),
                'Age Group': programme.get('target_audience', ''),
                'Class Start Date': datetime.strptime(programme.get('class_start', ''), "%a, %d %b %Y %H:%M:%S %Z").strftime('%Y-%m-%d') if programme.get('class_start', '') else '',
                'Class End Date': datetime.strptime(programme.get('class_end', ''), "%a, %d %b %Y %H:%M:%S %Z").strftime('%Y-%m-%d') if programme.get('class_end', '') else '',
                'Is Paid?': is_paid_value,
                'output': output_field
            })

        df = pd.DataFrame(export_data)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Application Details"

        # Append DataFrame rows to the worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Define the yellow fill for highlighting and highluighed if is_paid = Yes
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            is_paid = ws[f'E{i}'].value
            if is_paid == 'Yes':
                for cell in row:
                    cell.fill = yellow_fill

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"application_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@mod.route('/dashboard', methods=['GET'])
def get_dashboard_stats():
    try:
        total_students = Students.query.count()
        total_courses = Programme.query.count()
        total_forms = Application.query.count()
        
        return jsonify({
            'totalStudents': total_students,
            'totalCourses': total_courses,
            'totalForms': total_forms
        })
    except Exception as e:
        return make_response(jsonify({'error': str(e)}), 500)
    
